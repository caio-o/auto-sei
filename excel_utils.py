import os
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


def _build_excel_row(process_data: dict):
    if not isinstance(process_data, dict):
        process_data = {}

    numero_processo = str(process_data.get("numero_processo", "") or "").strip()

    return [
        datetime.now().strftime("%d/%m/%Y"),
        numero_processo,
        str(process_data.get("titular", "") or "").strip(),
        str(process_data.get("pa", "") or "").strip(),
        str(process_data.get("municipio", "") or "").strip(),
        str(process_data.get("numero_titulo", "") or "").strip(),
        str(process_data.get("gru", "") or "").strip(),
    ]


def build_excel_row_text(process_data: dict) -> str:
    return "\t".join(map(str, _build_excel_row(process_data)))


def export_process_to_excel(output_path: str, process_data: dict):
    if not isinstance(process_data, dict):
        process_data = {}

    pasta_destino = os.path.dirname(output_path)
    if pasta_destino:
        os.makedirs(pasta_destino, exist_ok=True)

    headers = [
        "Data de geração",
        "Número de processo",
        "Titular",
        "PA",
        "Município",
        "Número de títulos",
        "GRU",
    ]

    nova_linha = _build_excel_row(process_data)
    numero_processo = str(process_data.get("numero_processo", "") or "").strip()

    thin = Side(style="thin", color="D9D9D9")

    if os.path.exists(output_path):
        try:
            wb = load_workbook(output_path)
            ws = wb["Monitoramento"] if "Monitoramento" in wb.sheetnames else wb.active
        except PermissionError:
            raise PermissionError(
                f"Não foi possível abrir o arquivo porque ele está aberto:\n{output_path}"
            )
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Monitoramento"

        ws.append(headers)

        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        widths = {1: 18, 2: 24, 3: 32, 4: 18, 5: 24, 6: 20, 7: 18}
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = "A2"

    linha_existente = None
    if numero_processo:
        for row in range(2, ws.max_row + 1):
            valor = ws.cell(row=row, column=2).value
            if str(valor or "").strip() == numero_processo:
                linha_existente = row
                break

    if linha_existente:
        for col_idx, valor in enumerate(nova_linha, start=1):
            cell = ws.cell(row=linha_existente, column=col_idx)
            cell.value = valor
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    else:
        ws.append(nova_linha)
        ultima_linha = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=ultima_linha, column=col_idx)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    try:
        wb.save(output_path)
        wb.close()
    except PermissionError:
        raise PermissionError(
            f"Não foi possível salvar o arquivo porque ele está aberto:\n{output_path}"
        )
